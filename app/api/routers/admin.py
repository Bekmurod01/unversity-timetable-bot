from datetime import datetime
from io import BytesIO

from aiogram import Bot
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import admin_guard, get_db_session
from app.config import get_settings
from app.models import ExamDeadline, Teacher, TimetableLesson, UpdateLog
from app.schemas import BroadcastRequest, ExamCreate, TeacherUpsert, TimetableUploadRow
from app.services.notification_service import NotificationService
from app.services.timetable_service import TimetableService

router = APIRouter(prefix="/admin", tags=["admin"], dependencies=[Depends(admin_guard)])
settings = get_settings()


@router.get("/dashboard")
async def dashboard(db: AsyncSession = Depends(get_db_session)) -> dict:
    service = TimetableService(db)
    return await service.get_dashboard_stats()


@router.post("/timetable/upload-json")
async def upload_timetable_json(rows: list[TimetableUploadRow], db: AsyncSession = Depends(get_db_session)) -> dict:
    service = TimetableService(db)
    await service.replace_timetable([x.model_dump() for x in rows])
    return {"status": "ok", "rows": len(rows)}


@router.post("/timetable/upload-excel")
async def upload_timetable_excel(file: UploadFile = File(...), db: AsyncSession = Depends(get_db_session)) -> dict:
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only Excel files are supported")

    content = await file.read()
    workbook = load_workbook(filename=BytesIO(content), read_only=True)
    sheet = workbook.active
    all_rows = list(sheet.iter_rows(values_only=True))
    if not all_rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    headers = [str(x).strip().lower() if x is not None else "" for x in all_rows[0]]
    required = ["group_name", "subject", "teacher", "room", "day", "start_time", "end_time"]
    missing = [x for x in required if x not in headers]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(missing)}")

    rows: list[dict] = []
    for values in all_rows[1:]:
        if values is None:
            continue
        item = {headers[idx]: values[idx] for idx in range(min(len(headers), len(values)))}
        if not item.get("group_name") or not item.get("subject"):
            continue
        rows.append(
            {
                "group_name": str(item.get("group_name")).strip(),
                "subject": str(item.get("subject")).strip(),
                "teacher": str(item.get("teacher", "")).strip(),
                "room": str(item.get("room", "")).strip(),
                "day": str(item.get("day", "")).strip(),
                "start_time": str(item.get("start_time", "")).strip(),
                "end_time": str(item.get("end_time", "")).strip(),
                "status": str(item.get("status", "active")).strip() or "active",
            }
        )

    service = TimetableService(db)
    await service.replace_timetable(rows)
    return {"status": "ok", "rows": len(rows)}


@router.post("/timetable/manual")
async def manual_upsert_timetable(row: TimetableUploadRow, db: AsyncSession = Depends(get_db_session)) -> dict:
    lesson = TimetableLesson(
        group_name=row.group_name,
        subject=row.subject,
        teacher=row.teacher,
        room=row.room,
        day=row.day.lower(),
        start_time=datetime.strptime(row.start_time, "%H:%M").time(),
        end_time=datetime.strptime(row.end_time, "%H:%M").time(),
        status=row.status,
    )
    db.add(lesson)
    await db.commit()
    return {"status": "ok"}


@router.get("/teachers")
async def list_teachers(db: AsyncSession = Depends(get_db_session)) -> list[dict]:
    teachers = list((await db.execute(select(Teacher).order_by(Teacher.name))).scalars().all())
    return [{"id": t.id, "name": t.name, "subject": t.subject} for t in teachers]


@router.post("/teachers")
async def add_teacher(payload: TeacherUpsert, db: AsyncSession = Depends(get_db_session)) -> dict:
    db.add(Teacher(name=payload.name, subject=payload.subject))
    await db.commit()
    return {"status": "ok"}


@router.put("/teachers/{teacher_id}")
async def edit_teacher(teacher_id: int, payload: TeacherUpsert, db: AsyncSession = Depends(get_db_session)) -> dict:
    teacher = (await db.execute(select(Teacher).where(Teacher.id == teacher_id))).scalar_one_or_none()
    if not teacher:
        raise HTTPException(status_code=404, detail="Teacher not found")
    teacher.name = payload.name
    teacher.subject = payload.subject
    await db.commit()
    return {"status": "ok"}


@router.delete("/teachers/{teacher_id}")
async def delete_teacher(teacher_id: int, db: AsyncSession = Depends(get_db_session)) -> dict:
    await db.execute(delete(Teacher).where(Teacher.id == teacher_id))
    await db.commit()
    return {"status": "ok"}


@router.post("/broadcast")
async def broadcast(payload: BroadcastRequest, db: AsyncSession = Depends(get_db_session)) -> dict:
    bot = Bot(token=settings.bot_token)
    notifier = NotificationService(bot, db)
    sent = await notifier.broadcast(payload.message, payload.group_name, payload.year)
    await bot.session.close()
    return {"status": "ok", "sent": sent}


@router.post("/exams")
async def add_exam(payload: ExamCreate, db: AsyncSession = Depends(get_db_session)) -> dict:
    row = ExamDeadline(**payload.model_dump())
    db.add(row)
    await db.commit()
    return {"status": "ok", "id": row.id}


@router.get("/logs")
async def logs(db: AsyncSession = Depends(get_db_session)) -> list[dict]:
    rows = list((await db.execute(select(UpdateLog).order_by(UpdateLog.created_at.desc()).limit(200))).scalars().all())
    return [
        {
            "id": x.id,
            "group_name": x.group_name,
            "change_type": x.change_type,
            "details": x.details,
            "created_at": x.created_at.isoformat() if x.created_at else None,
        }
        for x in rows
    ]
